import os
from dotenv import load_dotenv

import openpyxl
import pandas as pd
import requests
from goose3 import Goose

from newsapi import NewsApiClient

# Load API Keys from .env file
load_dotenv()

# Set News API Key
newsapi = NewsApiClient(api_key=os.getenv("NEWS_API_KEY"))

'''
Fetches articles about the given company and saves them in the given filename as a Microsoft Excel file
'''
def create_company_articles_workbook(q='Nvidia market', filename='sample_fine_tune.xlsx', domains=None, excludeDomains=None):
	# Create Python Goose Article Extractor Object
	g = Goose()

	# Create a workbook object
	workbook = openpyxl.Workbook()

	# Select the active sheet
	sheet = workbook.active
	
	# Perform call for get_everything endpoint of News API
	news_articles = newsapi.get_everything(q=q, language='en', domains=domains, exclude_domains=excludeDomains) # News Article is a nested dictionary

	# These column titles are necessary for putting the xlsx in a format recognizable by the OpenAI CLI data preparation tool
	sheet["A1"] = "title"
	sheet["B1"] = "date_published"
	sheet["C1"] = "url"
	sheet["D1"] = "timeout"
	sheet["E1"] = "content"
	

	row = 2 # began at the next row

	# extract the article text from each news article, if URL request timesout write information to xlsx
	for news in news_articles.get('articles'):
		try :
			title = news.get('title')
			date = news.get('publishedAt')
			url = news.get('url')
			content = g.extract(url=url).cleaned_text
			timeout = False
		except requests.exceptions.ReadTimeout:
			title = news.get('title')
			date = news.get('publishedAt')
			url = news.get('url')
			content = news.get('content')
			timeout = True
		finally:
			print(url)
			sheet.cell(row=row, column=1, value = title);
			sheet.cell(row=row, column=2, value = date);
			sheet.cell(row=row, column=3, value = url);
			sheet.cell(row=row, column=4, value = timeout);
			sheet.cell(row=row, column=5, value = content);
			
			row += 1

	workbook.save(filename)
